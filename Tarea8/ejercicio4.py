
from openpyxl import load_workbook

libro_ventas=load_workbook(filename="escritura_datos.xlsx")
ws=libro_ventas.active

ws["C2"]=0.55

libro_ventas.save("mi_libro.xlsx")