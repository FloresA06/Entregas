
from openpyxl import Workbook

libro_d=Workbook()
ws=libro_d.active
ws.title="Ventas"

ws["A1"]="Producto"
ws["A2"]="Manzanas"
ws["A3"]="Naranjas"

ws["B1"]="Cantidad"
ws["B2"]=50
ws["B3"]=30

ws["C1"]="Precio"
ws["C2"]=0.5
ws["C3"]=0.75

libro_d.save("escritura_datos.xlsx")

