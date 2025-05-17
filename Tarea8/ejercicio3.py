
from openpyxl import load_workbook

libro_ventas=load_workbook(filename="escritura_datos.xlsx")
ws=libro_ventas.active

ventas=list()

productos=[ws["A2"].value,ws["A3"].value]
cantidad=[ws["B2"].value,ws["B3"].value]
precio=[ws["C2"].value,ws["C3"].value]

ventas.extend([productos,cantidad,precio])
print(ventas)

