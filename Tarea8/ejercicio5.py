
from openpyxl import load_workbook

libro_ventas=load_workbook(filename="mi_libro.xlsx")
ws=libro_ventas.active

ws["D1"]="Total"

ws["D2"]=ws["B2"].value*ws["C2"].value
ws["D3"]=ws["B3"].value*ws["C3"].value

libro_ventas.save("mi_libro2.xlsx")

print(ws["D2"].value)