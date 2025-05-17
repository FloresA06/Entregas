
from openpyxl import load_workbook

libro=load_workbook(filename="mi_libro.xlsx")
ws=libro.active
ws2=libro.create_sheet("Inventario")

ws2["A1"]="Producto"
ws2["A2"]="Manzanas"
ws2["A3"]="Naranjas"
ws2["B1"]="Stock"
ws2["B2"]=100
ws2["B3"]=80

libro.save("mi_libro3.xlsx")

inventario={
    ws2["A2"].value:{
    ws2["B1"].value:ws2["B2"].value
    },
    ws2["A3"].value:{
    ws2["B1"].value:ws2["B3"].value
    }
}

ventas={
    ws["A2"].value:{
    ws["B1"].value:ws["B2"].value,
    ws["C1"].value:ws["C2"].value
    },
    ws["A3"].value:{
    ws["B1"].value:ws["B3"].value,
    ws["C1"].value:ws["C3"].value
    }
}

print(ventas)
print(inventario)

